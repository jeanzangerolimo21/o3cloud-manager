from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
import unicodedata

from app.contatos.service import ContatoService
from app.core.pdf_assinatura import extrair_data_assinatura_pdf
from app.core.storage import StorageService
from app.integracoes.omie.client import OmieClient
from app.integracoes.omie.contrato_mapper import ContratoMapper
from app.parceiros.executivo_service import ParceiroExecutivoService
from app.parceiros.service import ParceiroService
from app.propostas.service import PropostaService
from app.repositories.cliente_repository import ClienteRepository
from app.repositories.contrato_repository import ContratoRepository


class ContratoService:
    STATUS_OPTIONS = {
        "RASCUNHO": "Rascunho",
        "ENVIADO_CLICKSIGN": "Enviado ClickSign",
        "AGUARDANDO_ASSINATURA": "Aguardando Assinatura",
        "ENCAMINHADO_PROJETO": "Encaminhado para Projeto",
        "CONCLUIDO": "Concluido",
        "ATIVO": "Contratos Ativos",
        "EM_IMPLANTACAO": "Em Implantacao",
        "SUSPENSO": "Suspenso",
        "ENCERRADO": "Encerrado",
        "CANCELADO": "Cancelado",
    }

    TIPO_VENDA_OPTIONS = {
        "USUARIO": "Usuario",
        "PROJETO": "Projeto",
    }

    @classmethod
    def sincronizar_contrato(cls, contrato_omie, vendedores_cache=None, projetos_cache=None, vinculos_cache=None):
        dados = ContratoMapper.from_omie(contrato_omie)
        cls._aplicar_nomes_omie(dados, vendedores_cache, projetos_cache)
        cls._aplicar_vinculos_comerciais_omie(dados, vinculos_cache)
        cliente = ClienteRepository.buscar_por_codigo_externo(dados["cliente_codigo_externo"])

        if not cliente:
            return {
                "status": "IGNORADO",
                "numero": dados["numero"],
                "motivo": "Cliente nao encontrado",
            }

        dados["cliente_id"] = cliente["id"]
        contrato = ContratoRepository.buscar_por_codigo_externo(dados["codigo_externo"])

        if not contrato:
            contrato = ContratoRepository.buscar_manual_por_numero(
                cliente["id"], dados.get("numero")
            )
        if not contrato:
            contrato = ContratoRepository.buscar_assinado_sem_codigo_por_cliente_valor(
                cliente["id"], dados.get("valor_mensal")
            )
        if not contrato:
            contrato = ContratoRepository.buscar_omie_ativo_por_cliente(cliente["id"])

        if contrato:
            ContratoRepository.atualizar_sync(contrato["id"], dados)
            duplicados = ContratoRepository.desativar_omie_ativos_por_cliente(
                cliente["id"],
                contrato["id"],
            )
            cls._registrar_historico_valor(contrato["id"], dados)
            return {
                "status": "UPDATE",
                "numero": dados["numero"],
                "duplicados_desativados": duplicados,
            }

        contrato_id = ContratoRepository.inserir(dados)
        cls._registrar_historico_valor(contrato_id, dados)
        return {"status": "INSERT", "numero": dados["numero"]}

    @classmethod
    def desativar_contratos_omie_ausentes(cls, codigos_externos):
        return ContratoRepository.desativar_omie_ativos_ausentes(codigos_externos)

    @staticmethod
    def _registrar_historico_valor(contrato_id, dados):
        from app.financeiro.reajuste_service import ReajusteContratoService
        ReajusteContratoService.registrar_historico_valor_se_necessario(contrato_id, dados, origem="OMIE")

    @classmethod
    def sincronizar(cls):
        omie = OmieClient()
        vendedores_cache = cls._indexar_cadastros_omie(omie.listar_vendedores)
        projetos_cache = cls._indexar_cadastros_omie(omie.listar_projetos)
        resposta = omie.listar_contratos()
        vinculos_cache = cls._cache_vinculos_comerciais_omie()
        return [
            cls.sincronizar_contrato(contrato, vendedores_cache, projetos_cache, vinculos_cache)
            for contrato in resposta.get("contratoCadastro", [])
        ]

    @classmethod
    def preencher_vinculos_comerciais_omie_existentes(cls):
        cache = cls._cache_vinculos_comerciais_omie()
        contratos = ContratoRepository.listar_omie_ativos_para_vinculos_comerciais()
        atualizados = 0
        parceiro_match = 0
        executivo_match = 0
        sem_parceiro = set()
        sem_executivo = set()

        for contrato in contratos:
            vinculos = cls._resolver_vinculos_comerciais_omie(
                contrato.get("vendedor_nome"),
                contrato.get("projeto_nome"),
                cache,
            )
            parceiro_id = vinculos.get("parceiro_id")
            executivo_id = vinculos.get("executivo_id")

            if parceiro_id:
                parceiro_match += 1
            elif contrato.get("vendedor_nome"):
                sem_parceiro.add(contrato.get("vendedor_nome"))

            if executivo_id:
                executivo_match += 1
            elif contrato.get("projeto_nome"):
                sem_executivo.add(contrato.get("projeto_nome"))

            if parceiro_id or executivo_id:
                ContratoRepository.atualizar_vinculos_comerciais_omie_sync(
                    contrato.get("id"),
                    parceiro_id=parceiro_id,
                    executivo_id=executivo_id,
                )
                atualizados += 1

        return {
            "processados": len(contratos),
            "atualizados": atualizados,
            "parceiro_match": parceiro_match,
            "executivo_match": executivo_match,
            "sem_parceiro": sorted(sem_parceiro),
            "sem_executivo": sorted(sem_executivo),
        }

    @classmethod
    def _aplicar_nomes_omie(cls, dados, vendedores_cache=None, projetos_cache=None):
        dados["vendedor_nome"] = cls._nome_cadastro_omie(
            dados.get("codigo_vendedor"),
            vendedores_cache,
        )
        dados["projeto_nome"] = cls._nome_cadastro_omie(
            dados.get("codigo_projeto"),
            projetos_cache,
        )

    @classmethod
    def _aplicar_vinculos_comerciais_omie(cls, dados, cache=None):
        cache = cache or cls._cache_vinculos_comerciais_omie()

        dados.update(
            cls._resolver_vinculos_comerciais_omie(
                dados.get("vendedor_nome"),
                dados.get("projeto_nome"),
                cache,
            )
        )

    @classmethod
    def _resolver_vinculos_comerciais_omie(cls, vendedor_nome, projeto_nome, cache):
        vendedor_chave = cls._normalizar_nome_vinculo(vendedor_nome)
        parceiro_nome = cache["vendedores_omie"].get(vendedor_chave)
        if not parceiro_nome and vendedor_chave:
            for chave_mapa, nome_mapa in cache["vendedores_omie"].items():
                if chave_mapa.startswith(f"{vendedor_chave} "):
                    parceiro_nome = nome_mapa
                    break
        parceiro = cache["parceiros"].get(
            cls._normalizar_nome_vinculo(parceiro_nome or vendedor_nome)
        )
        executivo = cache["executivos"].get(
            cls._normalizar_nome_vinculo(projeto_nome)
        )
        return {
            "parceiro_id": parceiro.get("id") if parceiro else None,
            "executivo_id": executivo.get("id") if executivo else None,
        }

    @classmethod
    def _cache_vinculos_comerciais_omie(cls):
        return {
            "vendedores_omie": cls._mapa_vendedores_omie(),
            "parceiros": cls._indexar_parceiros_por_nome(),
            "executivos": cls._indexar_executivos_por_nome(),
        }

    @classmethod
    def _mapa_vendedores_omie(cls):
        caminho = Path(__file__).resolve().parents[2] / "storage" / "temporarios" / "VENDEDORES.xlsx"
        if not caminho.exists():
            return {}

        try:
            from openpyxl import load_workbook
            workbook = load_workbook(caminho, read_only=True, data_only=True)
        except Exception:
            return {}

        try:
            worksheet = workbook.active
            mapa = {}
            for linha in worksheet.iter_rows(min_row=2, values_only=True):
                vendedor_omie = cls._texto_vinculo(linha[0] if len(linha) > 0 else None)
                parceiro_o3 = cls._texto_vinculo(linha[1] if len(linha) > 1 else None)
                if vendedor_omie and parceiro_o3:
                    mapa[cls._normalizar_nome_vinculo(vendedor_omie)] = parceiro_o3
            return mapa
        finally:
            workbook.close()

    @classmethod
    def _indexar_parceiros_por_nome(cls):
        indice = {}
        for parceiro in ParceiroService.listar_todos_ativos():
            for campo in ("nome_exibicao", "nome", "sigla"):
                chave = cls._normalizar_nome_vinculo(parceiro.get(campo))
                if chave and chave not in indice:
                    indice[chave] = parceiro
        return indice

    @classmethod
    def _indexar_executivos_por_nome(cls):
        indice = {}
        for executivo in ParceiroExecutivoService.listar_todos_ativos():
            chave = cls._normalizar_nome_vinculo(executivo.get("nome"))
            if chave and chave not in indice:
                indice[chave] = executivo
        return indice

    @staticmethod
    def _texto_vinculo(valor):
        if valor in (None, ""):
            return ""
        return str(valor).strip()

    @classmethod
    def _normalizar_nome_vinculo(cls, valor):
        texto = cls._texto_vinculo(valor)
        if not texto:
            return ""
        texto = unicodedata.normalize("NFKD", texto)
        texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
        texto = " ".join(texto.upper().replace(".", " ").replace("-", " ").split())
        return texto

    @staticmethod
    def _nome_cadastro_omie(codigo, cache):
        if not codigo:
            return None
        if not cache:
            return None
        return cache.get(int(codigo))

    @classmethod
    def _indexar_cadastros_omie(cls, listar):
        pagina = 1
        cadastros = {}

        while True:
            resposta = listar(pagina)
            for item in resposta.get("cadastro", []):
                codigo = item.get("codigo")
                if codigo:
                    cadastros[int(codigo)] = item.get("nome")

            total_paginas = resposta.get("total_de_paginas", pagina)
            if pagina >= total_paginas:
                break
            pagina += 1

        return cadastros

    @classmethod
    def listar(cls, filtros, pagina=1, limit=50):
        offset = (pagina - 1) * limit
        contratos = ContratoRepository.listar(limit=limit, offset=offset, **filtros)
        total = ContratoRepository.total(**filtros)
        return contratos, total, (total + limit - 1) // limit

    @classmethod
    def dashboard(cls, filtros):
        return ContratoRepository.dashboard(**filtros)

    @classmethod
    def contar_encaminhados_sem_arquivo(cls):
        return ContratoRepository.contar_encaminhados_sem_arquivo()

    @classmethod
    def diagnostico_pre_beta(cls, contrato, implantacao=None):
        implantacao = implantacao or {}
        itens = []

        def adicionar(tipo, titulo, detalhe, icone):
            itens.append({
                "tipo": tipo,
                "titulo": titulo,
                "detalhe": detalhe,
                "icone": icone,
                "classe": {
                    "ok": "success",
                    "fluxo": "secondary",
                    "pendencia": "warning",
                    "erro": "danger",
                }.get(tipo, "secondary"),
            })

        if contrato.get("proposta_id"):
            adicionar("ok", "Proposta vinculada", "Contrato possui rastreabilidade comercial por proposta.", "bi-link-45deg")
        else:
            adicionar("fluxo", "Contrato direto", "Fluxo valido para contratos diretos ou vindos de parceiro; nao exige proposta para implantacao.", "bi-signpost-2")

        if implantacao:
            adicionar("ok", "Implantacao vinculada", "Projeto operacional ja existe para este contrato.", "bi-hdd-network")
        elif contrato.get("status") in ("ENCAMINHADO_PROJETO", "EM_ELABORACAO", "EM_IMPLANTACAO"):
            adicionar("pendencia", "Implantacao pendente", "Contrato esta em fase operacional e ainda nao possui implantacao ativa.", "bi-kanban")

        if not contrato.get("cliente_cnpj"):
            adicionar("pendencia", "CNPJ do cliente pendente", "Comercial deve completar o cadastro antes da validacao Beta assistida.", "bi-person-vcard")
        if not (contrato.get("contato_email") or contrato.get("cliente_email")):
            adicionar("pendencia", "Contato sem email", "Email do contato ou do cliente deve ser revisado para comunicacoes de validacao.", "bi-envelope")
        if not contrato.get("executivo_id"):
            adicionar("pendencia", "Executivo nao informado", "Responsavel comercial deve ser preenchido quando houver dono do relacionamento.", "bi-person-badge")
        if not contrato.get("data_fechamento"):
            adicionar("pendencia", "Data de fechamento pendente", "Campo recomendado para recortes executivos e planejamento operacional.", "bi-calendar-event")
        if not contrato.get("valor_mensal") and not contrato.get("valor_promocional"):
            adicionar("pendencia", "Receita recorrente pendente", "Valor mensal sera base de conferencia antes da carga financeira oficial.", "bi-currency-dollar")
        if contrato.get("status") == "ATIVO" and not contrato.get("dia_faturamento"):
            adicionar("pendencia", "Dia de faturamento pendente", "Dado necessario para conferencia financeira durante a Beta.", "bi-receipt")

        if not itens:
            adicionar("ok", "Sem pendencias pre-Beta", "Contrato tem os dados basicos esperados para validacao assistida.", "bi-check2-circle")

        return itens

    @classmethod
    def contexto_form(cls):
        return {
            "clientes": ContratoRepository.listar_clientes_para_contrato(),
            "contatos": cls._contatos_para_form(),
            "propostas": cls._propostas_para_form(),
            "executivos": ParceiroExecutivoService.listar_todos_ativos(),
            "parceiros": ParceiroService.listar_todos_ativos(),
            "status_options": cls.STATUS_OPTIONS,
            "tipo_venda_options": cls.TIPO_VENDA_OPTIONS,
            "modelo_contrato": StorageService.caminho(StorageService.CONTRATOS, "Modelo_Contrato.pdf"),
        }

    @classmethod
    def _contatos_para_form(cls):
        return [
            {
                "id": contato.get("id"),
                "nome": contato.get("nome") or "",
                "empresa": contato.get("empresa") or "",
                "email": contato.get("email") or "",
                "telefone": contato.get("telefone") or contato.get("whatsapp") or "",
                "whatsapp": contato.get("whatsapp") or "",
            }
            for contato in ContatoService.listar_todos_ativos("REPRESENTANTE_LEGAL")
        ]

    @classmethod
    def _propostas_para_form(cls):
        propostas, _ = PropostaService.listar(ativo="1")
        return [
            cls._proposta_para_form(proposta.get("id"))
            for proposta in propostas
            if proposta.get("id")
        ]

    @classmethod
    def _proposta_para_form(cls, proposta_id):
        proposta = PropostaService.buscar_por_id(proposta_id)
        if not proposta:
            return {}
        valor_setup = (proposta.get("setup_ambiente_cloud") or Decimal("0.00")) + (proposta.get("instalacao_servidores") or Decimal("0.00"))
        quantidade_usuarios = sum(cls._inteiro(item.get("quantidade")) or 0 for item in proposta.get("licencas_items") or [])
        return {
            "id": proposta.get("id"),
            "codigo": proposta.get("codigo_proposta") or "",
            "titulo": proposta.get("titulo") or "",
            "cliente_nome": proposta.get("cliente_nome") or proposta.get("cliente_nome_fantasia") or proposta.get("cliente_razao_social") or "",
            "status": proposta.get("status") or "",
            "cliente_id": proposta.get("cliente_id"),
            "contato_id": proposta.get("contato_id"),
            "parceiro_id": proposta.get("parceiro_id"),
            "executivo_id": proposta.get("executivo_responsavel_id"),
            "contato_nome": proposta.get("contato_nome") or "",
            "contato_email": proposta.get("contato_email") or "",
            "contato_telefone": proposta.get("contato_telefone") or "",
            "valor_mensal": cls._string_decimal(proposta.get("total_mensal")),
            "valor_setup": cls._string_decimal(valor_setup),
            "valor_projeto": cls._string_decimal(proposta.get("parametrizacao_sistema")),
            "quantidade_usuarios": quantidade_usuarios or "",
            "observacoes": proposta.get("detalhes_negociacao") or proposta.get("condicoes_comerciais") or proposta.get("observacoes") or "",
        }

    @classmethod
    def criar(cls, dados, arquivo_preparado=None):
        dados = cls._normalizar(dados)
        cls._validar(dados, arquivo_preparado=arquivo_preparado, exigir_pdf=False)
        cls._aplicar_arquivo_preparado(dados, arquivo_preparado)
        contrato_id = ContratoRepository.inserir_manual(dados)
        from app.financeiro.reajuste_service import ReajusteContratoService
        ReajusteContratoService.registrar_historico_valor_se_necessario(contrato_id, dados, origem="MANUAL")
        return contrato_id

    @classmethod
    def atualizar(cls, contrato_id, dados, arquivo_preparado=None):
        contrato = ContratoRepository.buscar_por_id(contrato_id)
        if not contrato:
            raise ValueError("Contrato nao encontrado.")
        if contrato["origem"] != "MANUAL":
            raise ValueError("Contratos sincronizados do Omie so permitem edicao de vinculos comerciais.")

        dados = cls._normalizar(dados)
        cls._validar(dados, arquivo_preparado=arquivo_preparado, exigir_pdf=False)
        cls._aplicar_arquivo_preparado(dados, arquivo_preparado)
        ContratoRepository.atualizar(contrato_id, dados)
        from app.financeiro.reajuste_service import ReajusteContratoService
        ReajusteContratoService.registrar_historico_valor_se_necessario(contrato_id, dados, origem="MANUAL")

    @classmethod
    def atualizar_vinculos_comerciais(cls, contrato_id, dados):
        contrato = ContratoRepository.buscar_por_id(contrato_id)
        if not contrato:
            raise ValueError("Contrato nao encontrado.")
        if contrato["origem"] != "OMIE":
            raise ValueError("Esta edicao restrita e exclusiva para contratos do Omie.")

        dados = cls._normalizar_vinculos_comerciais(dados)
        ContratoRepository.atualizar_vinculos_comerciais(contrato_id, dados)

    @classmethod
    def salvar_assinado(cls, contrato_id, arquivo):
        contrato = ContratoRepository.buscar_por_id(contrato_id)
        if not contrato:
            raise ValueError("Contrato nao encontrado.")
        if not cls._arquivo_pdf(arquivo):
            raise ValueError("Envie um arquivo PDF assinado.")

        salvo = StorageService.salvar(arquivo, StorageService.CONTRATOS)
        data_assinatura = extrair_data_assinatura_pdf(StorageService.caminho(StorageService.CONTRATOS, salvo["nome"]))
        ContratoRepository.atualizar_arquivo_assinado(
            contrato_id,
            salvo["nome"],
            salvo["arquivo_original"],
            data_assinatura,
        )

    @classmethod
    def caminho_assinado(cls, contrato_id):
        contrato = ContratoRepository.buscar_por_id(contrato_id)
        if not contrato or not contrato.get("arquivo_assinado"):
            return None, None

        caminho = StorageService.caminho(StorageService.CONTRATOS, contrato["arquivo_assinado"])
        if not Path(caminho).exists():
            return None, None

        return caminho, contrato.get("arquivo_assinado_original") or contrato["arquivo_assinado"]

    @classmethod
    def _normalizar(cls, dados):
        dados = dict(dados)
        cliente = ClienteRepository.buscar_por_id(dados.get("cliente_id")) if dados.get("cliente_id") else None
        numero = (dados.get("numero") or "").strip()
        dados["numero"] = numero or "CTR-" + datetime.now().strftime("%Y%m%d%H%M%S")
        dados["descricao"] = (dados.get("descricao") or "").strip() or None
        dados["status"] = dados.get("status") or "RASCUNHO"
        dados["tipo_venda"] = dados.get("tipo_venda") or "USUARIO"
        dados["cliente"] = cliente

        for campo in (
            "cliente_id",
            "contato_id",
            "proposta_id",
            "executivo_id",
            "parceiro_id",
            "quantidade_usuarios",
            "dia_faturamento",
        ):
            dados[campo] = cls._inteiro(dados.get(campo))

        for campo in ("valor_mensal", "valor_setup", "valor_projeto", "valor_promocional"):
            dados[campo] = cls._decimal(dados.get(campo))

        for campo in (
            "inicio_vigencia",
            "fim_vigencia",
            "data_fechamento",
            "data_inicio_recorrencia",
            "data_ativacao",
        ):
            dados[campo] = (dados.get(campo) or "").strip() or None

        cls._preencher_dados_contato(dados)
        for campo in ("contato_nome", "contato_email", "contato_telefone", "observacoes"):
            dados[campo] = (dados.get(campo) or "").strip() or None

        return dados

    @classmethod
    def _normalizar_vinculos_comerciais(cls, dados):
        dados = dict(dados)
        for campo in ("contato_id", "executivo_id", "parceiro_id"):
            dados[campo] = cls._inteiro(dados.get(campo))
        cls._preencher_dados_contato(dados)
        for campo in ("contato_nome", "contato_email", "contato_telefone", "observacoes"):
            dados[campo] = (dados.get(campo) or "").strip() or None
        return dados

    @classmethod
    def _preencher_dados_contato(cls, dados):
        if not dados.get("contato_id"):
            return
        contato = ContatoService.buscar_por_id(dados.get("contato_id"))
        if not contato:
            return
        if contato.get("tipo_contato") != "REPRESENTANTE_LEGAL":
            raise ValueError("Selecione um contato do tipo Representante Legal para o contrato.")
        dados["contato_nome"] = dados.get("contato_nome") or contato.get("nome")
        dados["contato_email"] = dados.get("contato_email") or contato.get("email")
        dados["contato_telefone"] = dados.get("contato_telefone") or contato.get("telefone") or contato.get("whatsapp")

    @classmethod
    def _validar(cls, dados, arquivo_preparado=None, exigir_pdf=False):
        erros = []
        cliente = dados.get("cliente")

        if not cliente:
            erros.append("Informe a empresa/cliente")
        elif not cliente.get("cnpj"):
            erros.append("Informe o CNPJ")
        if not dados.get("contato_nome"):
            erros.append("Informe o nome do contato")
        if not dados.get("contato_email"):
            erros.append("Informe o e-mail do contato")
        if not dados.get("contato_telefone"):
            erros.append("Informe o telefone do contato")
        if not dados.get("data_fechamento"):
            erros.append("Informe a data de fechamento")

        for campo, label in (
            ("inicio_vigencia", "Data de inicio da vigencia"),
            ("fim_vigencia", "Data de fim da vigencia"),
            ("data_fechamento", "Data de fechamento"),
            ("data_inicio_recorrencia", "Data inicio recorrente"),
            ("data_ativacao", "Data de ativacao"),
        ):
            valor = dados.get(campo)
            if not valor:
                continue
            try:
                datetime.strptime(valor, "%Y-%m-%d")
            except ValueError:
                erros.append(f"{label} invalida. Use uma data valida no formato AAAA-MM-DD")

        if not dados.get("valor_mensal") or dados["valor_mensal"] <= 0:
            erros.append("Informe o valor recorrente")
        if dados.get("quantidade_usuarios") is not None and dados["quantidade_usuarios"] < 0:
            erros.append("A quantidade de usuarios nao pode ser negativa")
        if exigir_pdf and not cls._arquivo_pdf(arquivo_preparado):
            erros.append("Anexe o contrato em PDF")
        elif arquivo_preparado and arquivo_preparado.filename and not cls._arquivo_pdf(arquivo_preparado):
            erros.append("O contrato deve ser um arquivo PDF")
        if dados.get("status") not in cls.STATUS_OPTIONS:
            erros.append("Status invalido")
        if dados.get("tipo_venda") not in cls.TIPO_VENDA_OPTIONS:
            erros.append("Tipo de venda invalido")

        if erros:
            raise ValueError("|".join(erros))

    @classmethod
    def _aplicar_arquivo_preparado(cls, dados, arquivo):
        dados["arquivo_preparado"] = None
        dados["arquivo_preparado_original"] = None
        if arquivo and arquivo.filename:
            salvo = StorageService.salvar(arquivo, StorageService.CONTRATOS)
            dados["arquivo_preparado"] = salvo["nome"]
            dados["arquivo_preparado_original"] = salvo["arquivo_original"]

    @staticmethod
    def _arquivo_pdf(arquivo):
        return bool(arquivo and arquivo.filename and Path(arquivo.filename).suffix.lower() == ".pdf")


    @staticmethod
    def _string_decimal(valor):
        if valor in (None, ""):
            return ""
        try:
            return str(Decimal(str(valor)).quantize(Decimal("0.01")))
        except (InvalidOperation, ValueError):
            return ""

    @staticmethod
    def _inteiro(valor):
        if valor in (None, ""):
            return None
        try:
            return int(valor)
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _decimal(valor):
        if valor in (None, ""):
            return None
        texto = str(valor).strip().replace("R$", "").replace(".", "").replace(",", ".")
        try:
            return Decimal(texto)
        except (InvalidOperation, ValueError):
            return None
